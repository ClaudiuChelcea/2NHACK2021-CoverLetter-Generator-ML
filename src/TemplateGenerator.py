import docx
import shutil
import os
from pathlib import Path
from docx.shared import Pt
from datetime import date
from docx2pdf import convert
import sys
from openpyxl import load_workbook


def manageDocs(companyName, jobTitle):
    pathCompany = companyName.replace(" ", "")
    pathJobTitle = jobTitle.replace(" ", "")

    # make folder named company name
    newPath = os.path.join(os.getcwd(), pathCompany)
    Path(newPath).mkdir(parents=True, exist_ok=True)

    # copy template to the new folder and rename it to the name of the position
    templatePath = os.path.join(os.getcwd(), 'trainingData/cover-letter.docx')
    destination = os.path.join(newPath, pathJobTitle + '.docx')
    shutil.copy(templatePath, destination)
    return destination


def replace_string(filename, find, replace):
    """
    for a specified path to a file, finds all instances of a string and replaces it with desired string
    """
    doc = docx.Document(filename)
    style = doc.styles['Normal']
    font = style.font
    font.name = 'Garamond'
    font.size = Pt(12)
    for p in doc.paragraphs:
        if find in p.text:
            print('SEARCH FOUND!!')
            text = p.text.replace(find, replace)
            p.text = text
            p.style = doc.styles['Normal']
    doc.save(filename)


def convertToPDF(docxpath):
    convert(docxpath)

class XLrow:
    def __init__(self, companyName, positionName, jobID, contactName):
        self.companyName = companyName
        self.positionName = positionName
        self.jobID = jobID
        self.contactName = contactName

    def __str__(self):
        return str(self.companyName) + " " + str(self.positionName) + " " + str(self.jobID) + " " + str(
            self.contactName)


def readXL():
    XLrows = []
    workbook = load_workbook(filename="trainingData/Jobs.xlsx")
    sh = workbook.active
    for rowIdx in range(2, sh.max_row + 1):
        row = []
        for colIdx in range(1, sh.max_column + 1):
            cell_obj = sh.cell(row=rowIdx, column=colIdx)
            row.append(str(cell_obj.value))
        obj = XLrow(row[0], row[1], row[2], row[3])
        XLrows.append(obj)
    return XLrows


def batchGenerate(letterinfo):
    for row in letterInfo:
        companyName = row.companyName
        jobTitle = row.positionName
        jobId = row.jobID
        contactName = row.contactName

        if jobId == "d":
            jobId = ""
        else:
            jobId = " with job id " + jobId

        if contactName == "d":
            contactName = "Sir/Ma'am"

        replaceDict = {'#companyName#': companyName,
                       '#date#': date.today().strftime("%B %d, %Y"),
                       '#jobTitle#': jobTitle,
                       '#jobId#': jobId,
                       '#contactName#': contactName
                       }

        destination = manageDocs(companyName, jobTitle)

        for find in replaceDict:
            replace_string(destination, find, replaceDict[find])

        convertToPDF(destination)


if __name__ == "__main__":
    # change cwd to path of script
    os.chdir(os.path.dirname(sys.argv[0]))
    letterInfo = readXL()
    batchGenerate(letterInfo)
